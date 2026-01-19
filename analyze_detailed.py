import openpyxl
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('Sheets actual 19-01-26.xlsx', data_only=True)

print('='*100)
print('ANÁLISIS DETALLADO DEL SISTEMA ACTUAL')
print('='*100)

# HOJA 1: INVERSIÓN GASTOS
print('\n\n' + '='*100)
print('1. HOJA: INVERSIÓN GASTOS')
print('='*100)

ws1 = wb['INVERSION  GASTOS']
rows = list(ws1.iter_rows(values_only=True))

print('\nESTRUCTURA:')
print('- Esta hoja parece dividirse en DOS secciones paralelas:')
print('  * GASTOS TONY (columnas B-H)')
print('  * GASTOS FACU (columnas J-P)')
print('\n- Encabezados TONY:', rows[2][1:8])
print('- Encabezados FACU:', rows[2][9:16])

print('\n- Total filas con datos:', len(rows))
print(f'- Rango de fechas (estimado): {rows[3][1]} hasta aproximadamente últimas filas')

# Analizar resumen
print(f'\n- Totales encontrados en fila 2:')
print(f'  * TONY - USD: {rows[1][5]}, Pesos: {rows[1][6]}')
print(f'  * FACU - USD: {rows[1][13]}, Pesos: {rows[1][14]}')
print(f'  * Total combinado Pesos: {rows[2][18]}')
print(f'  * Total combinado USD: {rows[3][18]}')

print('\nRESUMEN INVERSORES (fila 6-7):')
print(f'  * FACU: {rows[6][18]} pesos, {rows[6][19]} USD, Promedio: {rows[6][20]}')
print(f'  * TONY: {rows[7][18]} pesos, {rows[7][19]} USD, Promedio: {rows[7][20]}')

# HOJA 2: STOCK Y VENTAS
print('\n\n' + '='*100)
print('2. HOJA: STOCK Y VENTAS')
print('='*100)

ws2 = wb['STOCK Y VENTAS']
rows2 = list(ws2.iter_rows(values_only=True))

print('\nESTRUCTURA:')
print('- Encabezados:', rows2[1][1:13])
print('- Total transacciones:', len(rows2) - 2)

# Analizar tipos de transacciones
tipos = {}
clientes = set()
for i, row in enumerate(rows2[2:], 3):
    if row[2]:  # DETALLE
        tipo = str(row[2]).strip()
        tipos[tipo] = tipos.get(tipo, 0) + 1
    if row[3]:  # CLIENTE
        clientes.add(str(row[3]).strip())

print(f'\n- Cantidad de clientes únicos: {len(clientes)}')
print(f'\n- Tipos de transacciones encontradas ({len(tipos)} tipos):')
for tipo, count in sorted(tipos.items(), key=lambda x: x[1], reverse=True)[:20]:
    print(f'  * {tipo}: {count} veces')

# HOJA 3: GASTOS FIJOS
print('\n\n' + '='*100)
print('3. HOJA: GASTOS FIJOS Y OTROS')
print('='*100)

ws3 = wb['GASTOS FIJOS Y OTROS']
rows3 = list(ws3.iter_rows(values_only=True))

print('\nESTRUCTURA:')
print('- Encabezados:', rows3[2][2:7])
print('- Gastos fijos mensuales:')

total_gastos_fijos = 0
for row in rows3[3:15]:
    if row[3] and row[4]:
        print(f'  * {row[3]}: ${row[4]:,.0f} - Vence: {row[5]}')
        if isinstance(row[4], (int, float)):
            total_gastos_fijos += row[4]

print(f'\nTOTAL GASTOS FIJOS MENSUALES: ${total_gastos_fijos:,.0f}')

print('\n\nINFORMACIÓN DE PRODUCCIÓN (columnas laterales):')
print(f'- Sala 1: {rows3[6][10]} g x cultivo, ciclo: {rows3[6][11]}')
print(f'- Sala 2: {rows3[7][10]} g x cultivo, ciclo: {rows3[7][11]}')
print(f'- Total: {rows3[8][10]} g, {rows3[8][11]} ciclos al año')
print(f'- Total gramos x año: {rows3[9][11]}')
print(f'- Precio estimado x mes: {rows3[11][10]}')
print(f'- Total precio estimado x año: {rows3[11][11]}')
print(f'- Diferencia: {rows3[12][11]}')
print(f'- Precio venta: ${rows3[13][11]:,.0f}')
print(f'- Saldo: ${rows3[14][11]:,.0f}')

wb.close()

print('\n\n' + '='*100)
print('FIN DEL ANÁLISIS')
print('='*100)
