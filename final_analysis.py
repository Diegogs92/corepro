import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('Sheets actual 19-01-26.xlsx', data_only=True)

print('='*100)
print('INTERPRETACIÓN DEL SISTEMA ACTUAL DE GOOGLE SHEETS')
print('='*100)

# HOJA 1
print('\n' + '='*100)
print('HOJA 1: INVERSIÓN GASTOS')
print('='*100)

ws1 = wb['INVERSION  GASTOS']
rows = [list(row) for row in ws1.iter_rows(values_only=True)]

print('\nINTERPRETACIÓN:')
print('Esta hoja registra las INVERSIONES iniciales del club, divididas por inversor:')
print('\n1. SECCIÓN IZQUIERDA - GASTOS TONY (Columnas B-H):')
print('   - FECHA: Cuándo se realizó el gasto')
print('   - DETALLE: Descripción del gasto')
print('   - MONTO USD: Valor en dólares')
print('   - MONTO $: Valor en pesos')
print('   - PRECIO DOLAR: Cotización del día')
print(f'   - TOTAL: USD {rows[1][5] if rows[1][5] else "N/A"} | $ {rows[1][6] if rows[1][6] else "N/A"}')

print('\n2. SECCIÓN DERECHA - GASTOS FACU (Columnas J-P):')
print('   - Misma estructura que Tony')
print(f'   - TOTAL: USD {rows[1][13] if rows[1][13] else "N/A"} | $ {rows[1][14] if rows[1][14] else "N/A"}')

print('\n3. RESUMEN GLOBAL (Columnas R-V):')
print(f'   - Total combinado: ${rows[2][18] if rows[2][18] else "N/A"}')
print(f'   - Total USD: {rows[3][18] if rows[3][18] else "N/A"}')

# Contar gastos
gastos_tony = sum(1 for row in rows[3:] if row[1] and isinstance(row[1], type(rows[3][1])))
gastos_facu = sum(1 for row in rows[3:] if row[9] and isinstance(row[9], type(rows[3][9])))

print(f'\n4. ESTADÍSTICAS:')
print(f'   - Gastos registrados Tony: ~{gastos_tony}')
print(f'   - Gastos registrados Facu: ~{gastos_facu}')

# HOJA 2
print('\n\n' + '='*100)
print('HOJA 2: STOCK Y VENTAS')
print('='*100)

ws2 = wb['STOCK Y VENTAS']
rows2 = [list(row) for row in ws2.iter_rows(values_only=True)]

print('\nINTERPRETACIÓN:')
print('Esta hoja es el CORE del negocio. Registra TODO el movimiento operativo:')
print('\nCOLUMNAS:')
print('  - FECHA: Cuándo ocurrió la transacción')
print('  - DETALLE: Tipo de operación (WEED, ESQUEJE, KIT, GASTOS, SUELDOS, etc.)')
print('  - CLIENTES: Quién compró o a quién se le pagó')
print('  - PRECIO VTA: Precio unitario')
print('  - ENTREGA: Quién entregó (TINO, FACU, TONY)')
print('  - MODO PAGO: EFECTIVO, TRANSFERENCIA, etc.')
print('  - CANTIDAD GR: Gramos vendidos (negativo = salida de stock)')
print('  - ESQUE/KITS: Cantidad de esquejes/kits vendidos')
print('  - GASTOS: Marcador si es un gasto (1.0)')
print('  - TOTAL: Monto total de la transacción')
print('  - DEUDAS: Saldo pendiente del cliente')
print('  - NOTAS: Observaciones')

# Análisis de datos
ventas_weed = 0
ventas_esqueje = 0
gastos = 0
total_ingresos = 0
total_egresos = 0

clientes_unicos = set()
tipos_operacion = {}

for row in rows2[2:]:
    detalle = str(row[2]).strip() if row[2] else ''
    cliente = str(row[3]).strip() if row[3] else ''
    total = row[10] if row[10] and isinstance(row[10], (int, float)) else 0

    if detalle:
        tipos_operacion[detalle] = tipos_operacion.get(detalle, 0) + 1

    if cliente:
        clientes_unicos.add(cliente)

    if 'WEED' in detalle:
        ventas_weed += 1
    elif 'ESQUEJE' in detalle or 'KIT' in detalle:
        ventas_esqueje += 1

    if row[9] == 1.0:  # Es un gasto
        gastos += 1
        total_egresos += abs(total)
    elif total > 0:
        total_ingresos += total

print(f'\nESTADÍSTICAS:')
print(f'  - Total transacciones: {len(rows2) - 2}')
print(f'  - Clientes únicos: {len(clientes_unicos)}')
print(f'  - Ventas de WEED: {ventas_weed}')
print(f'  - Ventas de ESQUEJES/KITS: {ventas_esqueje}')
print(f'  - Gastos registrados: {gastos}')
print(f'  - Ingresos totales estimados: ${total_ingresos:,.0f}')
print(f'  - Egresos totales estimados: ${total_egresos:,.0f}')

print(f'\nTOP 10 TIPOS DE OPERACIÓN:')
for tipo, count in sorted(tipos_operacion.items(), key=lambda x: x[1], reverse=True)[:10]:
    print(f'  {count:3d}x - {tipo}')

# HOJA 3
print('\n\n' + '='*100)
print('HOJA 3: GASTOS FIJOS Y OTROS')
print('='*100)

ws3 = wb['GASTOS FIJOS Y OTROS']
rows3 = [list(row) for row in ws3.iter_rows(values_only=True)]

print('\nINTERPRETACIÓN:')
print('Esta hoja contiene DOS secciones:')

print('\n1. SECCIÓN IZQUIERDA - GASTOS FIJOS MENSUALES:')
print('   Gastos recurrentes que el club debe pagar cada mes:\n')

total_fijos = 0
for i, row in enumerate(rows3[3:15]):
    if row[3] and row[4]:
        detalle = row[3]
        try:
            monto = float(row[4]) if row[4] else 0
            vence = row[5]
            nota = row[6] if row[6] else ''
            print(f'   {int(row[2]) if row[2] else i}. {detalle}: ${monto:,.0f} - Vence: {vence} {nota}')
            total_fijos += monto
        except:
            pass

print(f'\n   TOTAL GASTOS FIJOS MENSUALES: ${total_fijos:,.0f}')

print('\n2. SECCIÓN DERECHA - PROYECCIÓN DE PRODUCCIÓN:')
print('   Cálculos de capacidad productiva y rentabilidad:\n')

if rows3[6] and rows3[6][10]:
    print(f'   - Sala 1: {rows3[6][10]} g/cultivo, ciclo de {rows3[6][11]}')
if rows3[7] and rows3[7][10]:
    print(f'   - Sala 2: {rows3[7][10]} g/cultivo, ciclo de {rows3[7][11]}')
if rows3[8] and rows3[8][10]:
    print(f'   - Total producción: {rows3[8][10]} g')
    print(f'   - Ciclos al año: {rows3[8][11]}')
if rows3[9] and rows3[9][11]:
    print(f'   - Total gramos/año: {rows3[9][11]}')

wb.close()

print('\n\n' + '='*100)
print('CONCLUSIONES Y PROBLEMÁTICAS DETECTADAS')
print('='*100)

print('''
1. DATOS DISPERSOS: La información está dividida en múltiples hojas sin relación clara
   - Inversiones en una hoja separada
   - Operaciones diarias en otra
   - Gastos fijos en una tercera

2. DUPLICACIÓN DE ESFUERZO:
   - Los gastos aparecen tanto en "INVERSIÓN GASTOS" como en "STOCK Y VENTAS"
   - No queda claro qué gastos van en cada lado

3. FALTA DE ESTRUCTURA RELACIONAL:
   - No hay un sistema de IDs para relacionar entidades
   - Los clientes se repiten con variaciones ("COLO", "COLO ", etc.)
   - No hay categorización clara de productos

4. CONTROL DE STOCK DEFICIENTE:
   - El stock se calcula restando gramos en la columna "CANTIDAD GR"
   - No hay un inventario claro por variedad
   - No se pueden hacer consultas como "¿cuánto stock tengo de Gellato?"

5. MEZCLA DE CONCEPTOS:
   - Una misma hoja mezcla ventas, gastos, sueldos, compras
   - Dificulta hacer reportes específicos

6. SIN TRAZABILIDAD:
   - No se puede rastrear fácilmente el flujo de dinero
   - No hay balance automático
   - No se relacionan los gastos de inversión con la operación

7. CÁLCULOS MANUALES:
   - Los totales parecen calcularse con fórmulas de Excel
   - Propenso a errores
   - Difícil auditoría
''')

print('='*100)
