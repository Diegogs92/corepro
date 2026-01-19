"""
Script de migración de datos del Excel al nuevo modelo de datos
Convierte "Sheets actual 19-01-26.xlsx" al formato del nuevo modelo TypeScript
"""

import openpyxl
import json
import sys
from datetime import datetime
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

# Cargar el Excel
wb = openpyxl.load_workbook('Sheets actual 19-01-26.xlsx', data_only=True)

# ============================================================================
# FUNCIONES AUXILIARES
# ============================================================================

def clean_string(s):
    """Limpia y normaliza strings"""
    if s is None:
        return None
    return str(s).strip()

def normalize_cliente(nombre):
    """Normaliza nombres de clientes para evitar duplicados"""
    if not nombre:
        return None
    # Remover espacios extra, convertir a title case
    return clean_string(nombre).upper().strip()

def to_iso_date(date_obj):
    """Convierte datetime a string ISO"""
    if isinstance(date_obj, datetime):
        return date_obj.isoformat()
    return None

def generate_id(prefix, counter):
    """Genera ID único"""
    return f"{prefix}-{counter:04d}"

# ============================================================================
# MIGRACIÓN HOJA 1: INVERSIÓN GASTOS
# ============================================================================

print("=" * 100)
print("MIGRACIÓN HOJA 1: INVERSIÓN GASTOS")
print("=" * 100)

ws_inversion = wb['INVERSION  GASTOS']
rows_inversion = list(ws_inversion.iter_rows(values_only=True))

# Inversores (hardcoded basado en el análisis)
inversores = [
    {
        "id": "inv-facu",
        "nombre": "Facu",
        "montoInvertidoUSD": float(rows_inversion[1][13]) if rows_inversion[1][13] else 0,
        "montoInvertidoPesos": float(rows_inversion[1][14]) if rows_inversion[1][14] else 0,
        "porcentajeParticipacion": 58.0,
        "fechaIngreso": "2023-11-01T00:00:00",
        "activo": True
    },
    {
        "id": "inv-tony",
        "nombre": "Tony",
        "montoInvertidoUSD": float(rows_inversion[1][5]) if rows_inversion[1][5] else 0,
        "montoInvertidoPesos": float(rows_inversion[1][6]) if rows_inversion[1][6] else 0,
        "porcentajeParticipacion": 42.0,
        "fechaIngreso": "2023-11-01T00:00:00",
        "activo": True
    }
]

print(f"\nInversores creados: {len(inversores)}")

# Gastos de inversión
gastos_inversion = []
gi_counter = 1

# Gastos de Tony (columnas B-H, desde fila 4)
for i, row in enumerate(rows_inversion[3:], 4):
    if row[1] and isinstance(row[1], datetime):  # Si hay fecha en columna B
        gasto = {
            "id": generate_id("gi", gi_counter),
            "inversorId": "inv-tony",
            "fecha": to_iso_date(row[1]),
            "detalle": clean_string(row[2]) if row[2] else "",
            "categoria": "OTROS",  # Categorizar manualmente después
            "montoUSD": float(row[5]) if row[5] and isinstance(row[5], (int, float)) else 0,
            "montoPesos": float(row[6]) if row[6] and isinstance(row[6], (int, float)) else 0,
            "precioDolar": float(row[7]) if row[7] and isinstance(row[7], (int, float)) else 0
        }
        gastos_inversion.append(gasto)
        gi_counter += 1

# Gastos de Facu (columnas J-P, desde fila 4)
for i, row in enumerate(rows_inversion[3:], 4):
    if row[9] and isinstance(row[9], datetime):  # Si hay fecha en columna J
        gasto = {
            "id": generate_id("gi", gi_counter),
            "inversorId": "inv-facu",
            "fecha": to_iso_date(row[9]),
            "detalle": clean_string(row[10]) if row[10] else "",
            "categoria": "OTROS",
            "montoUSD": float(row[13]) if row[13] and isinstance(row[13], (int, float)) else 0,
            "montoPesos": float(row[14]) if row[14] and isinstance(row[14], (int, float)) else 0,
            "precioDolar": float(row[15]) if row[15] and isinstance(row[15], (int, float)) else 0
        }
        gastos_inversion.append(gasto)
        gi_counter += 1

print(f"Gastos de inversión migrados: {len(gastos_inversion)}")

# ============================================================================
# MIGRACIÓN HOJA 2: STOCK Y VENTAS
# ============================================================================

print("\n" + "=" * 100)
print("MIGRACIÓN HOJA 2: STOCK Y VENTAS")
print("=" * 100)

ws_stock = wb['STOCK Y VENTAS']
rows_stock = list(ws_stock.iter_rows(values_only=True))

# Analizar y separar datos
socios_dict = {}
ventas = []
gastos_operativos = []
items_venta = []

socio_counter = 1
venta_counter = 1
gasto_counter = 1
item_counter = 1

# Procesar cada fila (desde fila 3, índice 2)
for i, row in enumerate(rows_stock[2:], 3):
    fecha = row[1]
    detalle = clean_string(row[2])
    cliente = normalize_cliente(row[3])
    precio_vta = row[4]
    entregador = clean_string(row[5])
    modo_pago = clean_string(row[6])
    cantidad_gr = row[7]
    cantidad_esquejes = row[8]
    es_gasto = row[9]
    total = row[10]
    deudas = row[11]
    notas = clean_string(row[12])

    if not fecha or not isinstance(fecha, datetime):
        continue

    # Crear/actualizar socio si no existe
    if cliente and cliente not in socios_dict:
        socios_dict[cliente] = {
            "id": generate_id("socio", socio_counter),
            "nombre": cliente,
            "tipo": "CLIENTE_FRECUENTE",
            "fechaRegistro": to_iso_date(fecha),
            "activo": True,
            "saldo": 0,
            "limiteCredito": 50000
        }
        socio_counter += 1

    # Es un gasto?
    if es_gasto == 1.0:
        gasto = {
            "id": generate_id("gasto", gasto_counter),
            "numero": gasto_counter,
            "fecha": to_iso_date(fecha),
            "categoriaId": "cat-varios",  # Categorizar después
            "detalle": detalle or "Gasto varios",
            "proveedor": cliente,
            "monto": abs(float(total)) if total and isinstance(total, (int, float)) else 0,
            "metodoPago": modo_pago.upper() if modo_pago else "EFECTIVO",
            "pagado": True,
            "esRecurrente": False
        }
        gastos_operativos.append(gasto)
        gasto_counter += 1

    # Es una venta?
    elif detalle and any(word in detalle.upper() for word in ['WEED', 'ESQUEJE', 'KIT']) and total and total > 0:
        socio_id = socios_dict[cliente]["id"] if cliente and cliente in socios_dict else None

        venta = {
            "id": generate_id("venta", venta_counter),
            "numero": venta_counter,
            "fecha": to_iso_date(fecha),
            "socioId": socio_id,
            "vendedorId": None,  # Mapear después
            "subtotal": float(total) if total and isinstance(total, (int, float)) else 0,
            "descuento": 0,
            "total": float(total) if total and isinstance(total, (int, float)) else 0,
            "estadoPago": "PAGADO" if deudas == 0 else "PENDIENTE",
            "montoPagado": float(total) if deudas == 0 else 0,
            "saldoPendiente": float(deudas) if deudas and isinstance(deudas, (int, float)) else 0,
            "metodoPago": modo_pago.upper() if modo_pago else "EFECTIVO",
            "entregado": True,
            "fechaEntrega": to_iso_date(fecha),
            "notas": notas
        }
        ventas.append(venta)

        # Crear item de venta
        cantidad = 0
        precio_unitario = 0
        producto_tipo = "WEED"

        if cantidad_gr and cantidad_gr < 0:
            cantidad = abs(float(cantidad_gr))
            producto_tipo = "WEED"
            precio_unitario = precio_vta if precio_vta else 6000
        elif cantidad_esquejes and cantidad_esquejes < 0:
            cantidad = abs(float(cantidad_esquejes))
            producto_tipo = "ESQUEJE"
            precio_unitario = precio_vta if precio_vta else 10000

        if cantidad > 0:
            item = {
                "id": generate_id("item", item_counter),
                "ventaId": venta["id"],
                "productoId": f"prod-{producto_tipo.lower()}",
                "descripcion": detalle,
                "cantidad": cantidad,
                "precioUnitario": float(precio_unitario),
                "subtotal": float(total) if total else 0,
                "descuento": 0,
                "total": float(total) if total else 0
            }
            items_venta.append(item)
            item_counter += 1

        venta_counter += 1

socios = list(socios_dict.values())

print(f"\nSocios migrados: {len(socios)}")
print(f"Ventas migradas: {len(ventas)}")
print(f"Items de venta migrados: {len(items_venta)}")
print(f"Gastos operativos migrados: {len(gastos_operativos)}")

# ============================================================================
# MIGRACIÓN HOJA 3: GASTOS FIJOS
# ============================================================================

print("\n" + "=" * 100)
print("MIGRACIÓN HOJA 3: GASTOS FIJOS")
print("=" * 100)

ws_fijos = wb['GASTOS FIJOS Y OTROS']
rows_fijos = list(ws_fijos.iter_rows(values_only=True))

gastos_fijos = []
gf_counter = 1

for i, row in enumerate(rows_fijos[3:15], 4):
    if row[3] and row[4]:  # Si hay detalle y monto
        try:
            monto = float(row[4]) if isinstance(row[4], (int, float)) else 0
            if monto > 0:
                gasto = {
                    "id": generate_id("gfijo", gf_counter),
                    "numero": gf_counter,
                    "fecha": "2025-01-10T00:00:00",
                    "categoriaId": "cat-fijo",
                    "detalle": clean_string(row[3]),
                    "monto": monto,
                    "metodoPago": "TRANSFERENCIA",
                    "pagado": True,
                    "esRecurrente": True,
                    "frecuencia": "MENSUAL",
                    "notas": clean_string(row[6]) if row[6] else None
                }
                gastos_fijos.append(gasto)
                gf_counter += 1
        except:
            pass

print(f"Gastos fijos migrados: {len(gastos_fijos)}")

# ============================================================================
# GUARDAR RESULTADOS
# ============================================================================

print("\n" + "=" * 100)
print("GUARDANDO RESULTADOS")
print("=" * 100)

resultado = {
    "inversores": inversores,
    "gastosInversion": gastos_inversion,
    "socios": socios,
    "ventas": ventas,
    "itemsVenta": items_venta,
    "gastosOperativos": gastos_operativos,
    "gastosFijos": gastos_fijos,
    "estadisticas": {
        "totalInversores": len(inversores),
        "totalGastosInversion": len(gastos_inversion),
        "totalSocios": len(socios),
        "totalVentas": len(ventas),
        "totalItemsVenta": len(items_venta),
        "totalGastosOperativos": len(gastos_operativos),
        "totalGastosFijos": len(gastos_fijos)
    }
}

# Guardar JSON
with open('datos_migrados.json', 'w', encoding='utf-8') as f:
    json.dump(resultado, f, ensure_ascii=False, indent=2)

print(f"\nArchivo generado: datos_migrados.json")
print(f"\nESTADÍSTICAS FINALES:")
print(f"  - Inversores: {len(inversores)}")
print(f"  - Gastos de inversión: {len(gastos_inversion)}")
print(f"  - Socios: {len(socios)}")
print(f"  - Ventas: {len(ventas)}")
print(f"  - Items de venta: {len(items_venta)}")
print(f"  - Gastos operativos: {len(gastos_operativos)}")
print(f"  - Gastos fijos: {len(gastos_fijos)}")

print("\n" + "=" * 100)
print("MIGRACIÓN COMPLETADA")
print("=" * 100)

wb.close()
