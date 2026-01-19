import openpyxl
import sys

# Configurar encoding para Windows
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('Sheets actual 19-01-26.xlsx', data_only=True)

print('=' * 80)
print('ANÁLISIS DEL ARCHIVO EXCEL')
print('=' * 80)

for sheet_name in wb.sheetnames:
    print(f'\n{"="*80}')
    print(f'HOJA: {sheet_name}')
    print(f'{"="*80}')

    ws = wb[sheet_name]

    # Contar filas con datos
    rows_with_data = 0
    max_row_found = 0
    max_col_found = 0

    all_rows = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            rows_with_data += 1
            max_row_found += 1
            all_rows.append(row)
            if len([c for c in row if c is not None]) > 0:
                max_col_found = max(max_col_found, len(row))

    print(f'\nFilas con datos: {rows_with_data}')
    print(f'Columnas: {max_col_found}')

    if rows_with_data > 0:
        print(f'\n--- Primeras 15 filas ---')
        for i, row in enumerate(all_rows[:15], 1):
            print(f'Fila {i}: {row}')

        if rows_with_data > 15:
            print(f'\n... ({rows_with_data - 15} filas más) ...')

    print('')

wb.close()
